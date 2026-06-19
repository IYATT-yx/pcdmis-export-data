"""
file: dataprocessor.py
description: 数据处理模块。将从 PC-DMIS 导出的数据转写到 Excel 文件，并根据要求设置小数精度和超差颜色。
author: IYATT-yx
copyright:  Copyright (c) 2026 IYATT-yx.
            Licensed under the MIT License. See LICENSE file in the project root for full license information.
"""
from common import Common
import constants

import csv
import os
import hashlib
import datetime
import shutil
from enum import Enum
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

class MeasurementStatus(Enum):
    OK = 0,
    OverPlus = 1,
    UnderMinus = -1    

class HorizontalAlign(Enum):
    left = "left"
    center = "center"
    right = "right"

class VerticalAlign(Enum):
    top = "top"
    center = "center"
    bottom = "bottom"

class Colors(Enum):
    WHITE = "FFFFFF"          # 白色
    BLACK = "000000"          # 黑色
    RED = "FF0000"            # 红色
    GREEN = "00FF00"          # 绿色
    BLUE = "0000FF"           # 蓝色
    YELLOW = "FFFF00"         # 黄色
    CYAN = "00FFFF"           # 青色
    MAGENTA = "FF00FF"        # 品红（洋红色）
    PINK = "FFC0CB"           # 粉红色
    ORANGE = "FFA500"         # 橙色
    GRAY = "808080"           # 灰色
    SILVER = "C0C0C0"         # 银色
    LIGHT_GREEN = "90EE90"    # 浅绿色
    BROWN = "A52A2A"          # 棕色
    PURPLE = "800080"         # 紫色
    VIOLET = "EE82EE"         # 紫罗兰色
    OLIVE = "808000"          # 橄榄色
    NAVY_BLUE = "000080"      # 海军蓝
    TEAL = "008080"           # 水绿色
    LIME = "00FF00"           # 酸橙绿
    MAROON = "800000"         # 棕红色
    DARK_GREEN = "006400"     # 深绿色
    GOLD = "FFD700"           # 金色
    INDIGO = "4B0082"         # 蓝紫色
    CORAL = "FF7F50"          # 珊瑚色
    LAVENDER = "E6E6FA"       # 薰衣草色
    BEIGE = "F5F5DC"          # 米色
    MINT_CREAM = "F5FFFA"     # 薄荷奶油色
    LIGHT_GRAY = "D3D3D3"     # 浅灰色
    DARK_GRAY = "A9A9A9"      # 深灰色
    SKY_BLUE = "87CEEB"       # 天蓝色
    TURQUOISE = "40E0D0"      # 宝石绿
    TOMATO = "FF6347"         # 番茄红
    HOT_PINK = "FF69B4"       # 热粉红色
    SLATE_GRAY = "708090"     # 石板灰
    LIGHT_BLUE = "ADD8E6"     # 浅蓝色
    DARK_BLUE = "00008B"      # 深蓝色
    LIGHT_CORAL = "F08080"    # 浅珊瑚色
    DARK_CYAN = "008B8B"      # 深青色
    LIGHT_YELLOW = "FFFFE0"   # 浅黄色
    DARK_GOLDENROD = "B8860B" # 深金菊黄
    LIGHT_STEEL_BLUE = "B0C4DE" # 浅钢蓝色
    DARK_SLATE_GRAY = "2F4F4F" # 深石板灰
    LIGHT_SLATE_GRAY = "778899" # 浅石板灰
    DARK_TURQUOISE = "00CED1" # 深宝石绿
    LIGHT_SEA_GREEN = "20B2AA" # 浅海绿色
    DARK_ORCHID = "9932CC"    # 深兰花紫
    LIGHT_SALMON = "FFA07A"   # 浅鲑鱼色
    DARK_RED = "8B0000"       # 深红色
    LIGHT_GOLDENROD = "FAFA00" # 浅菊黄色
    DARK_MAGENTA = "8B008B"   # 深品红

def calculateCsvDigest(dataRows):
    sha256 = hashlib.sha256()
    for row in dataRows:
        if len(row) >= 12:
            keyColumns = row[0:9]
            keyColumns.append(row[11])
            rowStr = '|'.join(keyColumns)
            sha256.update(rowStr.encode('utf-8'))
    return sha256.hexdigest()

def cleanRawDataRows(rawDataRows: list) -> list:
    """
    对从 CSV 读取的原始数据行进行清洗与合并处理
    
    CSV 字段映射索引参考：
    0: ID, 1: Feature1, 2: Feature2, 3: Feature3, 4: AxisLetter, 5: Unit,
    6: Nominal, 7: Plus, 8: Minus, 9: Measured, 10: Bonus, 11: Type
    """
    cleanedRows = []
    currentShellId = ''
    currentShellFeatures = ['', '', '']

    for row in rawDataRows:
        if len(row) < 12:
            continue
            
        idVal = row[0].strip()
        feat1 = row[1].strip()
        axis = row[4].strip()
        unit = row[5].strip()
        
        # 提取并转换为字符串便于判断，避免因空格导致逻辑错误
        nominal = row[6].strip()
        plus = row[7].strip()
        minus = row[8].strip()
        measured = row[9].strip()

        # ---- 场景 1: 清除无实质数据的纯基准行 (如 ID=A) ----
        # 特征：有 ID，但没有特征元素名称，且数值多为 0 或空
        if idVal and not feat1 and (nominal in ('0', '0.0', '')) and (measured in ('0', '0.0', '')):
            continue

        # ---- 场景 2: 下属数据行（ID 为空，继承上一行外壳的 ID 和特征） ----
        if not idVal:
            if currentShellId:  # 如果当前存在有效的外壳缓存
                row[0] = currentShellId
                row[1] = currentShellFeatures[0]
                row[2] = currentShellFeatures[1]
                row[3] = currentShellFeatures[2]
                cleanedRows.append(row)
            continue  # 处理完毕，跳过

        # ---- 场景 3: 带有 ID 的正常行 或 外壳行 ----
        # 检查这行是不是一个缺乏实质测量数据的“外壳”（例如位置公差的父级命令行）
        # 或者是形位公差 FCF 中未填充实质测量数据的空外壳（如 FCF跳动3）
        isShell = not axis or (nominal in ('0', '0.0', '') and plus in ('0', '0.0', '') and minus in ('0', '0.0', '') and measured in ('0', '0.0', ''))
        
        if isShell:
            # 记录外壳信息，供后续的空 ID 行继承
            currentShellId = idVal
            currentShellFeatures = [row[1], row[2], row[3]]
            # 此时先不把它加入 cleanedRows，因为它的实质数据在下一行
        else:
            # 说明是一条独立且完整的形位公差或尺寸数据（如同轴度1），直接保留
            cleanedRows.append(row)
            # 清空外壳缓存，防止污染后面的数据
            currentShellId = ''
            currentShellFeatures = ['', '', '']
    return cleanedRows

def writeExcelCell(ws, row: int, col: int, value, verticalAlign: VerticalAlign = VerticalAlign.center, horizontalAlign: HorizontalAlign = HorizontalAlign.left, fillColor: Colors = None) -> None:
    """
    封装单元格写入函数，支持通过枚举设置对齐方式与背景颜色
    
    Args:
        ws: 目标 WorkSheet 对象
        row: 行索引 (从 1 开始)
        col: 列索引 (从 1 开始)
        value: 写入的内容
        verticalAlign: 垂直对齐方式枚举 (VerticalAlign)
        horizontalAlign: 水平对齐方式枚举 (HorizontalAlign)
        fillColor: 颜色枚举 (Colors)，为 None 时不填充
    """
    if value == '':
        value = '/'
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical=verticalAlign.value, horizontal=horizontalAlign.value)
    if fillColor:
        cell.fill = PatternFill(start_color=fillColor.value, end_color=fillColor.value, fill_type='solid')

def setColumnWidth(ws, col_idx: int, width: float) -> None:
    """
    通过数字列号设置 Excel 工作表的列宽
    
    Args:
        ws: openpyxl 的 WorkSheet 对象
        col_idx: 数字列号 (从 1 开始)
        width: 目标列宽值
    """
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

# 单独一个函数先对数据数组精确到指定小数位
def formatDataPrecision(dataList: list, decimalPlaces: int) -> None:
    """
    原地修改二维数据列表中指定字段的小数精度。

    Args:
        dataList (list): PC-DMIS 原始数据行列表。
        decimalPlaces (int): 保留的小数位数。
    """
    for row in dataList:
        row[6] = round(float(row[6]), decimalPlaces)
        row[7] = round(float(row[7]), decimalPlaces)
        row[8] = round(float(row[8]), decimalPlaces)
        row[9] = round(float(row[9]), decimalPlaces)
        row[10] = round(float(row[10]), decimalPlaces)

# 轴字母映射名称
# 参考：
#       * https://nexus.hexagon.com/documentationcenter/en-US/bundle/pcdmis-2026.1-core/page/19_dimen_topics/Axes_Drop_Down_list.htm
#       * https://files.ms.hexagonmi.com/public/docs/How%20To%20Docs/True%20Position%20Dimensions.doc
axisMap = {
    'X': 'X轴坐标',
    'Y': 'Y轴坐标',
    'Z': 'Z轴坐标',
    'TP': '位置度(总结果)',
    'DF': '被测要素尺寸',
    'D1': '第一基准要素尺寸',
    'DRF': '基准参考框架',
    'D': '直径',
    'A': '角度',
    'M': '量值',
    'R': '半径',
    'L': '长度',
    'H': '高度',
    'PR': '极半径',
    'PA': '极角度',
    'RT': '沿报告矢量方向的偏差',
    'S': '烟曲面(横向切线)矢量方向的偏差',
    'RS': '沿曲面报告方向的偏差',
    'PD': '圆的直径（垂直于销钉矢量）',
    'T': '沿逼近矢量方向的误差（专用于曲面上的点）',
    'FORM': '特征的综合形状尺寸（形状误差）'
}

def axisLetterToName(dataList: list) -> None:
    """
    轴字母转名称

    Args:
        dataList (list): PC-DMIS 原始数据行列表。
    """
    for row in dataList:
        rawAxis = row[4]
        row[4] = axisMap.get(rawAxis, rawAxis)

def convertPcdCsvToExcel(dataPath: str = '', csvFilePath: str = r'C:\Temp\PC-DMIS-TEMP.csv', decimalPlaces: int = 4, sheetName: str = '导出数据', noProg: bool = False):
    r"""
    将 PC-DMIS 中导出的原始 CSV 文件转换为 Excel 文件。

    Args:
        dataPath (str): 数据文件保存目录，默认为空，表示使用默认目录。
        csvFilePath (str): PC-DMIS 导出的 CSV 文件路径，默认为 'C:\Temp\PC-DMIS-TEMP.csv'。
        decimalPlaces (int): 测量数据保留的小数位数，默认为 4。
        sheetName (str): Excel 工作表名称，默认为 '导出数据'。
        noProg (bool): 是否不备份测量程序副本，默认为 False。
    """
    if not os.path.exists(csvFilePath):
        msg: str = f'CSV 数据源文件 {csvFilePath} 不存在。'
        raise FileNotFoundError(msg)

    rawDataRows = []
    with open(csvFilePath, 'r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        try:
            metadata = next(reader)
            headers = next(reader)
        except StopIteration:
            msg: str = '警告：CSV 文件内容不足（缺少元数据或表头）'
            print(msg)
            return
        
        for row in reader:
            if row and len(row) >= 12:
                rawDataRows.append(row)

    if not rawDataRows:
        msg: str = '警告：CSV 文件无有效数据'
        print(msg)
        return
    
    # 元素据
    # ------------------------------------------------------------------------
    versionString: str = metadata[0].strip() # PC-DMIS 版本
    progName: str = metadata[1].strip() # 测量程序名称
    progNameWithoutExt: str = Path(progName).stem # 无扩展名的测量程序名称
    fullProgName: str = metadata[2].strip() # 测量程序完整路径
    minusTolShowNeg: bool = (metadata[5].strip().lower() == 'true') # PC-DMIS 负公差显示负号设置值
    dataRows = cleanRawDataRows(rawDataRows)
    currentDigest: str = calculateCsvDigest(dataRows)

    # 序列号
    serialNumber: str = metadata[3].strip() # 测量程序初始序列号
    sn: str = metadata[4].strip() # SN 变量值
    SN = '未填序列号'
    if sn != '':
        SN = sn
    elif serialNumber != '':
        SN = serialNumber

    # 时间戳
    # ------------------------------------------------------------------------
    nowObj = datetime.datetime.now()
    currentDate: str = nowObj.strftime('%Y-%m-%d')
    currentTime: str = nowObj.strftime('%H:%M:%S')
    currentDataTime: str = nowObj.strftime('(%Y%m%d_%H%M%S)')

    # 数据文件保存目录
    if dataPath.strip() == '':
        dataPath = constants.Path.defaultDataPath
    excelDir = os.path.join(dataPath, progNameWithoutExt) # 工具可执行文件目录下的 data 目录下，对应测量程序名目录下
    excelFilename = f'{progNameWithoutExt}({versionString}).xlsx'
    excelFilePath = Common.longPath(os.path.join(excelDir, excelFilename))
    progBackupFilename = f'{progNameWithoutExt}({versionString})({currentDataTime})({SN}).PRG'
    progBackupPath = Common.longPath(os.path.join(excelDir, progBackupFilename))

    if not os.path.exists(excelDir):
        os.makedirs(excelDir, exist_ok=True)

    # 备份测量程序文件
    if not noProg:
        try:
            shutil.copy2(fullProgName, progBackupPath)
            Common.setFileReadOnly(progBackupPath)
        except:
            print('警告：测量程序文件备份失败')

    # 打开/创建表格
    # ------------------------------------------------------------------------
    if os.path.exists(excelFilePath):
        Common.setFileReadOnly(excelFilePath, False)
        wb = load_workbook(excelFilePath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetName

    # 变更校验
    # ------------------------------------------------------------------------
    currentRow = ws.max_row
    isModified = False
    if currentRow > 1:
        if ws.cell(row=currentRow, column=1).value != currentDigest:
            isModified = True
        currentRow += 1
    else:
        isModified = True

    # 格式化指定小数位
    formatDataPrecision(dataRows, decimalPlaces)
    # 转换轴成名称
    axisLetterToName(dataRows)

    # 写表头
    # ------------------------------------------------------------------------
    startColIdx = 5
    endColIdx = startColIdx + len(dataRows)
    if isModified:
        # 项目名
        endRow = currentRow + 9
        # writeExcelCell(ws, currentRow, 1, constants.Basic.projectName)
        # writeExcelCell(ws, currentRow + 1, 1, constants.Basic.version)
        # writeExcelCell(ws, currentRow + 3, 1, constants.Basic.repository)
        # writeExcelCell(ws, currentRow + 4, 1, '项目开源地址↑')
        writeExcelCell(ws, endRow - 1, 1, '校验码↓')
        writeExcelCell(ws, endRow - 1, 2, '日期↓')
        writeExcelCell(ws, endRow - 1, 3, '时间↓')

        writeExcelCell(ws, currentRow, 4, '标识符')
        writeExcelCell(ws, currentRow + 1, 4, '特征1')
        writeExcelCell(ws, currentRow + 2, 4, '特征2')
        writeExcelCell(ws, currentRow + 3, 4, '特征3')
        writeExcelCell(ws, currentRow + 4, 4, '轴')
        writeExcelCell(ws, currentRow + 5, 4, '单位')
        writeExcelCell(ws, currentRow + 6, 4, '理论值')
        writeExcelCell(ws, currentRow + 7, 4, '上极限偏差')
        writeExcelCell(ws, currentRow + 8, 4, '下极限偏差')

        # 检测项
        for colIdx in range(startColIdx, endColIdx):
            for rowIdx in range(currentRow, endRow):
                writeExcelCell(ws, rowIdx, colIdx, dataRows[colIdx - startColIdx][rowIdx - currentRow])
        currentRow = endRow

    # 设置列宽
    # ------------------------------------------------------------------------
    for colIdx in range(1, endColIdx):
        setColumnWidth(ws, colIdx, 12)

    # 写数据
    # ------------------------------------------------------------------------
    writeExcelCell(ws, currentRow, 1, currentDigest)
    writeExcelCell(ws, currentRow, 2, currentDate)
    writeExcelCell(ws, currentRow, 3, currentTime)
    # 写检测数据
    outOfTol = False
    for colIdx in range(startColIdx, endColIdx):
        nominal = dataRows[colIdx - startColIdx][6]
        plusTol = dataRows[colIdx - startColIdx][7]
        minusTol = dataRows[colIdx - startColIdx][8]
        bouns = dataRows[colIdx - startColIdx][10]
        measured = dataRows[colIdx - startColIdx][9]
        type = dataRows[colIdx - startColIdx][11].strip().upper()
        # 单项数据合格状态标识
        color = None
        if type == 'F':
            if measured >= plusTol + bouns:
                color = Colors.RED
                outOfTol = True
        else:
            # 处理负公差显示负号
            if type == 'D':
                minusTol = -minusTol
            elif type == 'FD' and not minusTolShowNeg:
                minusTol = -minusTol

            if measured >= nominal + plusTol and not (plusTol == 0 and minusTol == 0):
                color = Colors.RED
                outOfTol = True
            elif measured <= nominal + minusTol and not (plusTol == 0 and minusTol == 0):
                color = Colors.MAGENTA
                outOfTol = True
        writeExcelCell(ws, currentRow, colIdx, measured, fillColor=color)

    # 整体合格状态标识
    color = Colors.GREEN
    if outOfTol:
        color = Colors.YELLOW
    writeExcelCell(ws, currentRow, 4, SN, fillColor=color)

    # 保存 Excel 文件
    wb.save(excelFilePath)
    Common.setFileReadOnly(excelFilePath)

    # 清除临时文件
    if os.path.exists(csvFilePath):
        os.remove(csvFilePath)