from commontools import CommonTools
from dialog import Dialog
from customexception import CustomException
from colors import Colors
import constants
from topmessagebox import TopMessagebox

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
from pcdmistools import PcdmisTools
import tkinter as tk

Dialog()

class ExcelTools:
    workBook = None
    sheet = None
    filePath = None
    currentRow = None

    @staticmethod
    def openExcel(filePath: str):
        """
        打开 Excel 文件。
        存在就直接打开，不存在就创建。

        Params:
            filePath: Excel 文件路径
        """
        if CommonTools.checkFileExist(filePath):
            Dialog.log(f'{filePath} 已存在，加载文件')
            CommonTools.setFileReadOnly(filePath, False) # 打开文件前先解除只读
            ExcelTools.workBook = openpyxl.load_workbook(filePath)
        else:
            Dialog.log(f'{filePath} 不存在，创建新文件')
            ExcelTools.workBook = openpyxl.Workbook()
        ExcelTools.sheet = ExcelTools.workBook.active
        ExcelTools.filePath = filePath
        ExcelTools.currentRow =  1 if ExcelTools.sheet.max_row == 1 else ExcelTools.sheet.max_row + 1

    @staticmethod
    def getDigestFromExcel() -> str:
        """
        获取最新行摘要.
        表格第一列用于存放上一次写入测量数据的测量项目特征摘要，本函数用于读取上次的摘要。

        Returns:
            str: 成功返回摘要内容，表格未填写摘要返回 None
        """
        if ExcelTools.sheet is None:
            raise CustomException('请打开 Excel 后进行操作', CustomException.CRITICAL)
        if ExcelTools.currentRow== 1:
            return None
        return ExcelTools.sheet.cell(ExcelTools.currentRow - 1, 1).value
    
    @staticmethod
    def setColWidth(col: int, width: int):
        ExcelTools.sheet.column_dimensions[get_column_letter(col)].width = width

    def setCellWrap(row: int, col: int):
        ExcelTools.sheet.cell(row, col).alignment = Alignment(wrap_text=True)

    @staticmethod
    def setCellPrecision(cell, precision: int):
        cell.number_format = '0.' + '0' * precision
    
    @staticmethod
    def writeHeader(dataList: list[dict]):
        """
        写表头

        Params:
            dataList: 测量数据列表
        """
        endRow = ExcelTools.currentRow + PcdmisTools.dataLen - 3
        endCol = len(dataList) + 5
        for col in range(1, endCol):
            for row in range(ExcelTools.currentRow, endRow):
                if col == 2:
                    ExcelTools.sheet.cell(row, col, PcdmisTools.dataKeys[row - ExcelTools.currentRow])
                    ExcelTools.setColWidth(col, 20)
                    continue
                elif col == 3 or col == 4:
                    ExcelTools.setColWidth(col, 12)
                elif col > 4:
                    value = dataList[col - 5][PcdmisTools.dataKeys[row - ExcelTools.currentRow]]
                    if value == 0:
                        value = ''
                    cell = ExcelTools.sheet.cell(
                        row,
                        col,
                        value
                    )
                    ExcelTools.setCellPrecision(cell, constants.Data.precision)
                    ExcelTools.setColWidth(col, 13)
                    ExcelTools.setCellWrap(row, col)
        ExcelTools.currentRow = endRow

    @staticmethod
    def writeData(digest: str, serialNumber: str, dataList: list[dict]):
        """
        写数据

        Params:
            digest: 测量项目字符串摘要
            serialNumber: 测量报告序列号
            dataList: 测量数据列表
        """
        endCol = len(dataList) + 5
        for col in range(1, endCol):
            if col == 1:
                ExcelTools.sheet.cell(ExcelTools.currentRow, col, digest)
            elif col == 2:
                ExcelTools.sheet.cell(ExcelTools.currentRow, col, serialNumber)
                ExcelTools.setCellWrap(ExcelTools.currentRow, col)
            elif col == 3:
                ExcelTools.sheet.cell(ExcelTools.currentRow, col, CommonTools.getTimeStamp(2))
            elif col == 4:
                ExcelTools.sheet.cell(ExcelTools.currentRow, col, CommonTools.getTimeStamp(3))
            else:
                data = dataList[col - 5]
                nominal = data['标称值']
                plus = data['上公差']
                minus = data['下公差']
                measured = data['实测值']
                bonus = data['补偿值']
                dataType = data['类型']

                cell = ExcelTools.sheet.cell(
                    ExcelTools.currentRow,
                    col,
                    measured
                )

                ExcelTools.setCellPrecision(cell, constants.Data.precision)

                if dataType == PcdmisTools.dataType.FCF:
                    if measured > nominal + plus + bonus:
                        ExcelTools.fillCellWithColor(ExcelTools.currentRow, col, constants.Data.overPlusColor)
                else:
                    if plus >= minus:
                        upper = nominal + plus
                        lower = nominal + minus
                    else:
                        upper = nominal + minus
                        lower = nominal + plus

                    if plus != minus:
                        if measured >= upper:
                            ExcelTools.fillCellWithColor(ExcelTools.currentRow, col, constants.Data.overPlusColor)
                        elif measured <= lower:
                            ExcelTools.fillCellWithColor(ExcelTools.currentRow, col, constants.Data.underMinusColor)

        ExcelTools.currentRow += 1

    @staticmethod
    def fillCellWithColor(row: int, col: int, color: Colors):
        """
        给单元格填充背景色

        Params:
            row: 行号
            col: 列号
            color: 颜色
        """
        ExcelTools.sheet.cell(row, col).fill = PatternFill(start_color=color.value, end_color=color.value, fill_type='solid')

    @staticmethod
    def write(serialNumber: str, dataList: list[dict]):
        """
        写数据到 Excel 文件

        Params:
            serialNumber: 测量报告序列号
            dataList: 测量数据列表
        """
        if len(dataList) == 0:
            raise CustomException('测量数据为空', CustomException.ERROR)
        
        # 实际记录的摘要为：测量项目字符串摘要 + 本工具版本号
        digest = PcdmisTools.calcDigest(dataList) + '_' + constants.Basic.version
        Dialog.log(f'测量项目的综合特征摘要：{digest}')
        if digest != ExcelTools.getDigestFromExcel():
            ExcelTools.writeHeader(dataList)
            Dialog.log('写表头')

        ExcelTools.writeData(digest, serialNumber, dataList)
        Dialog.log('写数据')

        ExcelTools.workBook.save(ExcelTools.filePath)
        CommonTools.setFileReadOnly(ExcelTools.filePath) # 设置文件只读
